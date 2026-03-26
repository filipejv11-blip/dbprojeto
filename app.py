"""
DB Diagnóstico v3.0 — Interface Principal
Novidades: Wizard de primeiro uso, entrada única inteligente,
           tempo real, checks customizados, notificações, Excel, tendência.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import threading, json, os, hashlib, time
from datetime import datetime
from typing import Optional, List, Dict, Any
import matplotlib; matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from engine import ConexaoConfig, CheckResult, DiagnosticoSession, DBConnector, CheckRunner, Scheduler
from engine import Store as SimpleStore   # store simples, sem criptografia
from smart_input import SmartInputDetector
from notifications import NotifConfig, NotificationDispatcher
from advanced_checks import RealtimeChecker, CustomCheck, CustomCheckRunner, DEFAULT_CUSTOM_CHECKS

# ── Dirs ─────────────────────────────────────────────────────────────────────
DATA_DIR = os.path.join(os.path.expanduser("~"), ".db_diagnostico")
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH  = os.path.join(DATA_DIR, "data.db")

# ── Tema ─────────────────────────────────────────────────────────────────────
THEMES = {
    "dark":  {"bg":"#0C1016","surface":"#131B24","surface2":"#1A2535","border":"#263244",
               "accent":"#38BDF8","accent2":"#0284C7","danger":"#F87171","warning":"#FBBF24",
               "ok":"#34D399","info":"#818CF8","text":"#E2E8F0","text2":"#94A3B8",
               "muted":"#475569","panel":"#0D1520","tree_sel":"#1E3A5F",
               "chart_bg":"#131B24","chart_fg":"#E2E8F0","chart_grid":"#1E2D40"},
    "light": {"bg":"#F1F5F9","surface":"#FFFFFF","surface2":"#E8EEF7","border":"#CBD5E1",
               "accent":"#0284C7","accent2":"#0369A1","danger":"#DC2626","warning":"#D97706",
               "ok":"#059669","info":"#4F46E5","text":"#0F172A","text2":"#334155",
               "muted":"#64748B","panel":"#F8FAFC","tree_sel":"#DBEAFE",
               "chart_bg":"#FFFFFF","chart_fg":"#0F172A","chart_grid":"#E2E8F0"},
}
C = THEMES["dark"]
ICONS = {"OK":"✅","ERRO":"❌","AVISO":"⚠️","INFO":"ℹ️"}
FM="Consolas"; FS="Segoe UI"

# ════════════════════════════════════════════════════════════════════════════
# APP PRINCIPAL
# ════════════════════════════════════════════════════════════════════════════
class DBDiagApp(ctk.CTk):

    def __init__(self):
        super().__init__()
        self._theme_name  = "dark"
        self._resultados: List[CheckResult] = []
        self._custom_checks: List[CustomCheck] = list(DEFAULT_CUSTOM_CHECKS)
        self._store       = SimpleStore(DB_PATH)
        self._scheduler   = Scheduler(self._store)
        self._notif_cfg   = NotifConfig()
        self._notif       = NotificationDispatcher(self._notif_cfg)
        self._rt_connector: Optional[DBConnector] = None
        self._rt_active   = False

        self.title("DB Diagnóstico v3.0")
        self.geometry("1340x840")
        self.minsize(1100, 700)
        self._apply_theme()
        self._build_ui()
        self._reload_profiles()

    # ── Theme ────────────────────────────────────────────────────────────────
    def _apply_theme(self):
        global C; C=THEMES[self._theme_name]
        ctk.set_appearance_mode("dark" if self._theme_name=="dark" else "light")
        self.configure(fg_color=C["bg"])

    def _toggle_theme(self):
        self._theme_name="light" if self._theme_name=="dark" else "dark"
        self._apply_theme()
        # FIX: para recursos ativos antes de destruir os widgets
        if self._rt_active: self._rt_active=False
        if self._rt_connector:
            try: self._rt_connector.close()
            except: pass
            self._rt_connector=None
        for w in self.winfo_children(): w.destroy()
        self._build_ui(); self._reload_profiles()
        # reinicia o guard após rebuild
        
    # ════════════════════════════════════════════════════════════════════════
    # UI SHELL
    # ════════════════════════════════════════════════════════════════════════
    def _build_ui(self):
        self.grid_columnconfigure(1,weight=1); self.grid_rowconfigure(0,weight=1)
        self._build_sidebar()
        self._build_main()
        self._nav("diag")   # só aqui — depois de sidebar E main estarem prontos

    def _build_sidebar(self):
        sb=ctk.CTkFrame(self,fg_color=C["surface"],corner_radius=0,width=300)
        sb.grid(row=0,column=0,sticky="nsew"); sb.grid_propagate(False); sb.grid_columnconfigure(0,weight=1)
        hdr=ctk.CTkFrame(sb,fg_color=C["surface2"],corner_radius=0,height=66)
        hdr.grid(row=0,column=0,sticky="ew"); hdr.grid_propagate(False)
        ctk.CTkLabel(hdr,text="⬡  DB Diagnóstico",font=ctk.CTkFont(FM,16,"bold"),text_color=C["accent"]).place(x=16,y=12)
        ctk.CTkLabel(hdr,text="v3.0",font=ctk.CTkFont(FS,10),text_color=C["muted"]).place(x=16,y=38)
        ctk.CTkButton(hdr,text="☀" if self._theme_name=="dark" else "🌙",width=30,height=30,corner_radius=15,fg_color=C["surface"],hover_color=C["border"],font=ctk.CTkFont(size=15),command=self._toggle_theme).place(relx=1.0,x=-12,y=18)
        nav=ctk.CTkFrame(sb,fg_color="transparent"); nav.grid(row=1,column=0,sticky="ew",padx=10,pady=8)
        self._nav_btns={}
        for label,key in [
            ("🔌  Diagnóstico",  "diag"),
            ("⚡  Tempo Real",   "realtime"),
            ("📊  Dashboard",    "dashboard"),
            ("📋  Resultados",   "resultados"),
            ("🕒  Histórico",    "historico"),
            ("🧩  Checks Custom","custom"),
            ("⚙️   Agendamento", "agendamento"),
            ("🔔  Notificações", "notificacoes"),
        ]:
            b=ctk.CTkButton(nav,text=label,anchor="w",font=ctk.CTkFont(FS,12),fg_color="transparent",hover_color=C["surface2"],text_color=C["text2"],height=36,corner_radius=8,command=lambda k=key:self._nav(k))
            b.pack(fill="x",pady=1); self._nav_btns[key]=b
        ctk.CTkFrame(sb,height=1,fg_color=C["border"]).grid(row=2,column=0,padx=10,sticky="ew")
        ctk.CTkLabel(sb,text="PERFIS SALVOS",font=ctk.CTkFont(FM,10,"bold"),text_color=C["muted"]).grid(row=3,column=0,padx=16,pady=(10,4),sticky="w")
        self.profiles_frame=ctk.CTkScrollableFrame(sb,fg_color="transparent",height=140)
        self.profiles_frame.grid(row=4,column=0,padx=6,sticky="ew")
        ctk.CTkFrame(sb,height=1,fg_color=C["border"]).grid(row=5,column=0,padx=10,sticky="ew")
        sb.grid_rowconfigure(6,weight=1)
        self.lbl_status=ctk.CTkLabel(sb,text="Pronto.",font=ctk.CTkFont(FM,11),text_color=C["muted"],wraplength=260,anchor="w")
        self.lbl_status.grid(row=6,column=0,padx=16,pady=8,sticky="sw")

    def _build_main(self):
        self.main=ctk.CTkFrame(self,fg_color=C["bg"],corner_radius=0)
        self.main.grid(row=0,column=1,sticky="nsew"); self.main.grid_rowconfigure(0,weight=1); self.main.grid_columnconfigure(0,weight=1)
        self.pages={}
        for key in ["diag","realtime","dashboard","resultados","historico","custom","agendamento","notificacoes"]:
            f=ctk.CTkFrame(self.main,fg_color=C["bg"],corner_radius=0); f.grid(row=0,column=0,sticky="nsew"); self.pages[key]=f
        self._build_page_diag(); self._build_page_realtime(); self._build_page_dashboard()
        self._build_page_resultados(); self._build_page_historico(); self._build_page_custom()
        self._build_page_agendamento(); self._build_page_notificacoes()

    def _nav(self,key):
        for k,b in self._nav_btns.items(): b.configure(fg_color=C["accent2"] if k==key else "transparent",text_color=C["text"] if k==key else C["text2"])
        self.pages[key].tkraise()
        if key=="historico": self._reload_history()
        
    def _phdr(self,parent,title,row=0):
        f=ctk.CTkFrame(parent,fg_color=C["surface2"],corner_radius=0,height=52)
        f.grid(row=row,column=0,sticky="ew"); f.grid_propagate(False)
        ctk.CTkLabel(f,text=title,font=ctk.CTkFont(FM,14,"bold"),text_color=C["text"]).place(x=20,y=14)

    def _card(self,parent,title,row,colspan=1):
        w=ctk.CTkFrame(parent,fg_color=C["surface"],corner_radius=12)
        w.grid(row=row,column=0,columnspan=colspan,padx=18,pady=8,sticky="ew"); w.grid_columnconfigure(0,weight=1)
        ctk.CTkLabel(w,text=title,font=ctk.CTkFont(FS,13,"bold"),text_color=C["text"]).grid(row=0,column=0,columnspan=4,padx=16,pady=(12,4),sticky="w")
        ctk.CTkFrame(w,height=1,fg_color=C["border"]).grid(row=1,column=0,columnspan=4,padx=12,sticky="ew")
        return w

    # ════════════════════════════════════════════════════════════════════════
    # PAGE: DIAGNÓSTICO (entrada única inteligente)
    # ════════════════════════════════════════════════════════════════════════
    def _build_page_diag(self):
        p=self.pages["diag"]; p.grid_columnconfigure(0,weight=1); p.grid_rowconfigure(1,weight=1)
        self._phdr(p,"🔌 Diagnóstico")
        scroll=ctk.CTkScrollableFrame(p,fg_color=C["bg"]); scroll.grid(row=1,column=0,sticky="nsew"); scroll.grid_columnconfigure(0,weight=1)

        # ── Entrada única inteligente ────────────────────────────────────────
        ic=self._card(scroll,"🧠 Entrada Inteligente de Conexão",row=0)
        ic.grid_columnconfigure(0,weight=1)
        ctk.CTkLabel(ic,text="Cole qualquer coisa: DSN, host, caminho de arquivo, string ADO.NET...",
                     font=ctk.CTkFont(FS,11),text_color=C["muted"]).grid(row=2,column=0,padx=16,pady=(4,2),sticky="w")
        ef=ctk.CTkFrame(ic,fg_color="transparent"); ef.grid(row=3,column=0,padx=14,pady=(2,4),sticky="ew"); ef.grid_columnconfigure(0,weight=1)
        self.smart_entry=ctk.CTkEntry(ef,placeholder_text="postgresql://user:pass@host/db  ou  localhost:3306  ou  C:\\banco.db  ou  .env",
                                       height=46,fg_color=C["panel"],border_color=C["border"],font=ctk.CTkFont(FM,13))
        self.smart_entry.grid(row=0,column=0,sticky="ew")
        self.smart_entry.bind("<KeyRelease>",self._smart_detect)
        self.smart_entry.bind("<FocusOut>",self._smart_detect)
        self._smart_badge=ctk.CTkLabel(ic,text="",font=ctk.CTkFont(FS,11),text_color=C["muted"])
        self._smart_badge.grid(row=4,column=0,padx=16,pady=(0,4),sticky="w")

        # campos extras (nome, senha)
        xf=ctk.CTkFrame(ic,fg_color="transparent"); xf.grid(row=5,column=0,padx=14,pady=(0,8),sticky="ew")
        xf.grid_columnconfigure((0,1,2),weight=1)
        ctk.CTkLabel(xf,text="Nome do perfil",font=ctk.CTkFont(FS,11),text_color=C["muted"]).grid(row=0,column=0,sticky="w",padx=4)
        ctk.CTkLabel(xf,text="Senha (se não estiver na string)",font=ctk.CTkFont(FS,11),text_color=C["muted"]).grid(row=0,column=1,sticky="w",padx=4)
        self.diag_nome=ctk.CTkEntry(xf,placeholder_text="ex: Produção",height=36,fg_color=C["panel"],border_color=C["border"],font=ctk.CTkFont(FM,12))
        self.diag_nome.grid(row=1,column=0,padx=4,sticky="ew")
        self.diag_senha=ctk.CTkEntry(xf,placeholder_text="••••••••",show="●",height=36,fg_color=C["panel"],border_color=C["border"],font=ctk.CTkFont(FM,12))
        self.diag_senha.grid(row=1,column=1,padx=4,sticky="ew")
        ctk.CTkButton(xf,text="💾 Salvar Perfil",height=36,width=140,fg_color=C["surface2"],hover_color=C["border"],border_width=1,border_color=C["border"],font=ctk.CTkFont(FS,11),command=self._salvar_perfil).grid(row=1,column=2,padx=4)

        # exemplos
        ex_f=ctk.CTkFrame(ic,fg_color=C["panel"],corner_radius=8); ex_f.grid(row=6,column=0,padx=14,pady=(0,12),sticky="ew")
        ctk.CTkLabel(ex_f,text="Exemplos (clique para usar):",font=ctk.CTkFont(FS,10,"bold"),text_color=C["muted"]).pack(anchor="w",padx=10,pady=(6,2))
        exs=ctk.CTkFrame(ex_f,fg_color="transparent"); exs.pack(fill="x",padx=6,pady=(0,6))
        for ex in SmartInputDetector.EXEMPLOS[:5]:
            ctk.CTkButton(exs,text=ex[:52],width=0,height=24,anchor="w",fg_color="transparent",hover_color=C["border"],text_color=C["muted"],font=ctk.CTkFont(FM,10),command=lambda e=ex:self._use_example(e)).pack(anchor="w",padx=2)

        # ── Checks ───────────────────────────────────────────────────────────
        ck=self._card(scroll,"Verificações",row=1); ck.grid_columnconfigure((0,1,2),weight=1)
        self._check_vars={}
        for i,(key,label) in enumerate(CheckRunner.CHECKS_AVAILABLE.items()):
            v=tk.BooleanVar(value=True); self._check_vars[key]=v
            ctk.CTkCheckBox(ck,text=label,variable=v,font=ctk.CTkFont(FS,12),text_color=C["text2"],fg_color=C["accent2"],checkmark_color=C["text"]).grid(row=i//3+2,column=i%3,padx=16,pady=8,sticky="w")
        # custom checks toggle
        self._use_custom=tk.BooleanVar(value=True)
        ctk.CTkCheckBox(ck,text="Checks SQL Customizados",variable=self._use_custom,font=ctk.CTkFont(FS,12),text_color=C["info"],fg_color=C["info"],checkmark_color=C["text"]).grid(row=4,column=0,padx=16,pady=8,sticky="w")

        # ── Ações ────────────────────────────────────────────────────────────
        af=ctk.CTkFrame(scroll,fg_color="transparent"); af.grid(row=2,column=0,pady=12)
        self.btn_run=ctk.CTkButton(af,text="▶  Executar Diagnóstico",font=ctk.CTkFont(FM,14,"bold"),fg_color=C["accent2"],hover_color="#0255A3",height=50,width=256,corner_radius=10,command=self._iniciar)
        self.btn_run.pack(side="left",padx=(0,8))
        ctk.CTkButton(af,text="💾 Excel",font=ctk.CTkFont(FS,12),fg_color=C["surface"],hover_color=C["border"],border_width=1,border_color=C["border"],height=50,width=90,corner_radius=10,command=self._exportar_excel).pack(side="left",padx=(0,4))
        ctk.CTkButton(af,text="🌐 HTML",font=ctk.CTkFont(FS,12),fg_color=C["surface"],hover_color=C["border"],border_width=1,border_color=C["border"],height=50,width=90,corner_radius=10,command=self._exportar_html_dlg).pack(side="left")
        self.progress=ctk.CTkProgressBar(scroll,height=6,fg_color=C["border"],progress_color=C["accent"])
        self.progress.grid(row=3,column=0,sticky="ew",padx=32,pady=(0,8)); self.progress.set(0)

    def _use_example(self,ex):
        self.smart_entry.delete(0,"end"); self.smart_entry.insert(0,ex); self._smart_detect()

    def _smart_detect(self,_=None):
        text=self.smart_entry.get().strip()
        if not text: self._smart_badge.configure(text=""); return
        cfg,conf,tipo,dica=SmartInputDetector.detect(text)
        if conf>=0.7: col=C["ok"]; icon="✅"
        elif conf>=0.4: col=C["warning"]; icon="⚠️"
        else: col=C["muted"]; icon="❓"
        self._smart_badge.configure(text=f"{icon}  {dica}  (confiança {int(conf*100)}%)",text_color=col)

    def _coletar_config_smart(self) -> Optional[ConexaoConfig]:
        text=self.smart_entry.get().strip()
        if not text: messagebox.showwarning("Conexão","Cole uma string de conexão no campo."); return None
        cfg,conf,tipo_det,dica=SmartInputDetector.detect(text)
        if not cfg or conf<0.3:
            messagebox.showerror("Entrada inválida",f"Não foi possível identificar a conexão.\n{dica}"); return None
        # FIX: para detecções parciais (host simples, host:porta) avisa o usuário
        # que campos obrigatórios podem estar faltando, em vez de tentar conectar e falhar
        if conf < 0.7 and cfg.tipo not in ("sqlite",):
            if not cfg.banco:
                messagebox.showwarning("Conexão incompleta",
                    f"Detectado: {dica}\n\nO nome do banco de dados não foi informado.\n"
                    "Use uma DSN completa ou preencha os campos manualmente.")
                return None
        senha=self.diag_senha.get()
        if senha: cfg.senha=senha
        nome=self.diag_nome.get().strip()
        if nome: cfg.nome=nome
        elif not cfg.nome: cfg.nome=cfg.tipo
        return cfg

    def _salvar_perfil(self):
        cfg=self._coletar_config_smart()
        if cfg:
            self._store.save_profile(cfg); self._reload_profiles()
            self._set_status(f"✅ Perfil '{cfg.nome}' salvo")

    def _reload_profiles(self):
        for w in self.profiles_frame.winfo_children(): w.destroy()
        profiles=self._store.list_profiles()
        if not profiles:
            ctk.CTkLabel(self.profiles_frame,text="Nenhum perfil",font=ctk.CTkFont(FS,11),text_color=C["muted"]).pack(padx=8,pady=4); return
        icons={"sqlite":"🗂","mysql":"🐬","postgresql":"🐘","sqlserver":"🔷"}
        for p in profiles:
            row=ctk.CTkFrame(self.profiles_frame,fg_color=C["surface2"],corner_radius=8,height=34); row.pack(fill="x",padx=4,pady=2); row.pack_propagate(False)
            ctk.CTkLabel(row,text=f"{icons.get(p.tipo,'🛢')}  {p.nome or p.tipo}",font=ctk.CTkFont(FS,11),text_color=C["text2"]).place(x=8,y=8)
            for w in [row]+row.winfo_children(): w.bind("<Button-1>",lambda e,cfg=p:self._load_profile(cfg))

    def _load_profile(self,cfg):
        full=self._store.list_profiles() and next((p for p in self._store.list_profiles() if p.nome==cfg.nome), cfg)
        if full: cfg=full
        # FIX: nunca expõe senha na string visível — coloca só no campo de senha mascarado
        if cfg.tipo != "sqlite":
            dsn=f"{cfg.tipo}://{cfg.usuario}@{cfg.host}{':'+str(cfg.porta) if cfg.porta else ''}/{cfg.banco}"
        else:
            dsn=cfg.arquivo or ""
        self.smart_entry.delete(0,"end"); self.smart_entry.insert(0,dsn)
        self.diag_nome.delete(0,"end"); self.diag_nome.insert(0,cfg.nome)
        if cfg.senha:
            self.diag_senha.delete(0,"end"); self.diag_senha.insert(0,cfg.senha)
        self._smart_detect(); self._nav("diag")

    def _iniciar(self):
        cfg=self._coletar_config_smart()
        if not cfg: return
        checks=[k for k,v in self._check_vars.items() if v.get()]
        if not checks: messagebox.showwarning("Atenção","Selecione ao menos uma verificação."); return
        self.btn_run.configure(state="disabled",text="⏳  Executando..."); self.progress.set(0); self._resultados=[]
        def _run():
            connector=DBConnector(cfg); ok,msg=connector.connect()
            if not ok: self.after(0,lambda:self._run_error(msg)); return
            runner=CheckRunner(connector)
            results=runner.run(checks,progress_cb=lambda p,m:self.after(0,lambda pct=p,ms=m:self._upd_prog(pct,ms)))
            if self._use_custom.get() and self._custom_checks:
                results+=CustomCheckRunner(connector).run(self._custom_checks)
            connector.close()
            self.after(0,lambda:self._run_done(results,cfg))
        threading.Thread(target=_run,daemon=True).start()

    def _upd_prog(self,pct,msg): self.progress.set(pct); self._set_status(msg)

    def _run_error(self,msg):
        self.btn_run.configure(state="normal",text="▶  Executar Diagnóstico"); self._set_status(f"❌ {msg}"); messagebox.showerror("Erro",msg)

    def _run_done(self,results,cfg):
        self._resultados=results; self.btn_run.configure(state="normal",text="▶  Executar Diagnóstico"); self.progress.set(1.0)
        now=datetime.now(); sid=hashlib.md5(now.isoformat().encode()).hexdigest()[:12]
        er=sum(1 for r in results if r.status=="ERRO"); av=sum(1 for r in results if r.status=="AVISO"); ok=sum(1 for r in results if r.status=="OK"); inf=sum(1 for r in results if r.status=="INFO")
        self._store.save_session(DiagnosticoSession(id=sid,profile_name=cfg.nome,db_type=cfg.tipo,started_at=now.isoformat(),finished_at=datetime.now().isoformat(),total=len(results),erros=er,avisos=av,ok=ok,info=inf,resultados=results))
        if self._notif:
            level="ERRO" if er else ("AVISO" if av else "OK")
            self._notif.dispatch(f"Diagnóstico: {cfg.nome}",f"{len(results)} checks — {er} erros, {av} avisos",level)
        self._refresh_resultados(); self._refresh_dashboard(); self._set_status(f"✅ {len(results)} checks | ❌ {er} | ⚠️ {av}"); self._nav("dashboard")

    # ════════════════════════════════════════════════════════════════════════
    # PAGE: TEMPO REAL
    # ════════════════════════════════════════════════════════════════════════
    def _build_page_realtime(self):
        p=self.pages["realtime"]; p.grid_columnconfigure(0,weight=1); p.grid_rowconfigure(1,weight=1)
        self._phdr(p,"⚡ Monitoramento em Tempo Real")
        top=ctk.CTkFrame(p,fg_color=C["surface2"],height=48,corner_radius=0); top.grid(row=1,column=0,sticky="ew"); top.grid_propagate(False)
        self.rt_status_lbl=ctk.CTkLabel(top,text="Desconectado",font=ctk.CTkFont(FM,11),text_color=C["muted"]); self.rt_status_lbl.pack(side="left",padx=16,pady=14)
        self.btn_rt_start=ctk.CTkButton(top,text="▶ Conectar",width=120,height=32,fg_color=C["ok"],hover_color="#059669",font=ctk.CTkFont(FS,12),command=self._rt_start); self.btn_rt_start.pack(side="right",padx=8,pady=8)
        self.btn_rt_stop=ctk.CTkButton(top,text="⏹ Parar",width=100,height=32,fg_color=C["danger"],hover_color="#B91C1C",font=ctk.CTkFont(FS,12),state="disabled",command=self._rt_stop); self.btn_rt_stop.pack(side="right",padx=4,pady=8)
        ctk.CTkLabel(top,text="Threshold queries lentas (s):",font=ctk.CTkFont(FS,10),text_color=C["muted"]).pack(side="right",padx=(0,4),pady=16)
        self.rt_threshold=ctk.CTkEntry(top,width=50,height=30,fg_color=C["panel"],border_color=C["border"],font=ctk.CTkFont(FM,11)); self.rt_threshold.pack(side="right",padx=4,pady=10); self.rt_threshold.insert(0,"5")

        nb=ctk.CTkFrame(p,fg_color=C["bg"]); nb.grid(row=2,column=0,sticky="nsew"); nb.grid_rowconfigure(0,weight=1); nb.grid_columnconfigure(0,weight=1); p.grid_rowconfigure(2,weight=1)
        self.rt_tabs=ctk.CTkTabview(nb,fg_color=C["panel"],segmented_button_fg_color=C["surface2"],segmented_button_selected_color=C["accent2"],segmented_button_unselected_color=C["surface2"],text_color=C["text2"])
        self.rt_tabs.pack(fill="both",expand=True,padx=12,pady=8)
        for tab in ["🔒 Locks & Sessões","🐌 Queries Lentas","📡 Conexões"]: self.rt_tabs.add(tab)

        style=ttk.Style(); style.theme_use("clam")
        style.configure("RT.Treeview",background=C["panel"],foreground=C["text"],fieldbackground=C["panel"],rowheight=26,font=(FM,10),relief="flat")
        style.configure("RT.Treeview.Heading",background=C["surface2"],foreground=C["muted"],font=(FS,10,"bold"),relief="flat")
        style.map("RT.Treeview",background=[("selected",C["tree_sel"])],foreground=[("selected",C["text"])])

        def _tree(parent,cols):
            f=ctk.CTkFrame(parent,fg_color=C["panel"],corner_radius=0); f.pack(fill="both",expand=True); f.grid_rowconfigure(0,weight=1); f.grid_columnconfigure(0,weight=1)
            t=ttk.Treeview(f,style="RT.Treeview",columns=[c[0] for c in cols],show="headings")
            for cid,hdr,w in cols: t.heading(cid,text=hdr,anchor="w"); t.column(cid,width=w)
            sb=ctk.CTkScrollbar(f,command=t.yview); t.configure(yscrollcommand=sb.set)
            t.grid(row=0,column=0,sticky="nsew"); sb.grid(row=0,column=1,sticky="ns")
            return t

        self.rt_locks_tree=_tree(self.rt_tabs.tab("🔒 Locks & Sessões"),[("pid","PID",60),("usuario","Usuário",120),("estado","Estado",100),("tipo_espera","Espera",120),("segundos","Seg.",60),("query","Query",400)])
        self.rt_slow_tree=_tree(self.rt_tabs.tab("🐌 Queries Lentas"),[("pid","PID",60),("usuario","Usuário",120),("duracao_seg","Duração (s)",90),("estado","Estado",90),("query","Query",500)])
        conn_tab=self.rt_tabs.tab("📡 Conexões")
        self.rt_conn_labels={}
        cf=ctk.CTkFrame(conn_tab,fg_color="transparent"); cf.pack(pady=32)
        for i,(k,label) in enumerate([("total","Conexões Totais"),("ativas","Ativas"),("idle","Idle"),("maximo","Máximo Permitido")]):
            kf=ctk.CTkFrame(cf,fg_color=C["surface"],corner_radius=12,width=160,height=90); kf.grid(row=0,column=i,padx=10); kf.grid_propagate(False)
            ctk.CTkLabel(kf,text=label,font=ctk.CTkFont(FS,11),text_color=C["muted"]).place(relx=0.5,y=14,anchor="n")
            lbl=ctk.CTkLabel(kf,text="—",font=ctk.CTkFont(FM,32,"bold"),text_color=C["accent"]); lbl.place(relx=0.5,y=40,anchor="n")
            self.rt_conn_labels[k]=lbl

    def _rt_start(self):
        cfg=self._coletar_config_smart()
        if not cfg: return
        self._rt_connector=DBConnector(cfg); ok,msg=self._rt_connector.connect()
        if not ok: messagebox.showerror("Tempo Real",msg); return
        self._rt_active=True; self.btn_rt_start.configure(state="disabled"); self.btn_rt_stop.configure(state="normal")
        self.rt_status_lbl.configure(text=f"Conectado: {cfg.tipo}/{cfg.nome}",text_color=C["ok"])
        # FIX: roda em thread daemon para não bloquear a UI
        threading.Thread(target=self._rt_loop_thread, daemon=True).start()

    def _rt_stop(self):
        self._rt_active=False
        if self._rt_connector: self._rt_connector.close(); self._rt_connector=None
        self.btn_rt_start.configure(state="normal"); self.btn_rt_stop.configure(state="disabled")
        self.rt_status_lbl.configure(text="Desconectado",text_color=C["muted"])

    def _rt_loop_thread(self):
        """Polling em thread daemon — nunca bloqueia o event loop do Tkinter."""
        while self._rt_active:
            try:
                if self._rt_connector is None: break
                chk=RealtimeChecker(self._rt_connector)
                threshold=int(self.rt_threshold.get() or 5)
                locks=chk.get_locks(); slow=chk.get_slow_queries(threshold); conns=chk.get_connections_summary()
                # FIX: captura local das variáveis para evitar closure bug
                self.after(0, lambda l=locks,s=slow,c=conns: self._rt_update(l,s,c))
            except Exception as e:
                err=str(e)
                self.after(0, lambda m=err: self.rt_status_lbl.configure(text=f"Erro: {m}",text_color=C["danger"]))
                break
            time.sleep(3)

    def _rt_update(self,locks,slow,conns):
        for item in self.rt_locks_tree.get_children(): self.rt_locks_tree.delete(item)
        for r in locks:
            vals=(r.get("pid",""),r.get("usuario",""),r.get("estado",""),r.get("tipo_espera",""),r.get("segundos",""),r.get("query","")[:120])
            self.rt_locks_tree.insert("","end",values=vals)
        for item in self.rt_slow_tree.get_children(): self.rt_slow_tree.delete(item)
        for r in slow:
            vals=(r.get("pid",""),r.get("usuario",""),r.get("duracao_seg",""),r.get("estado",""),r.get("query","")[:150])
            self.rt_slow_tree.insert("","end",values=vals)
        for k,lbl in self.rt_conn_labels.items(): lbl.configure(text=str(conns.get(k,"—")))

    # ════════════════════════════════════════════════════════════════════════
    # PAGE: CUSTOM CHECKS
    # ════════════════════════════════════════════════════════════════════════
    def _build_page_custom(self):
        p=self.pages["custom"]; p.grid_columnconfigure(0,weight=1); p.grid_rowconfigure(1,weight=1)
        self._phdr(p,"🧩 Checks SQL Customizados")
        main=ctk.CTkFrame(p,fg_color=C["bg"]); main.grid(row=1,column=0,sticky="nsew"); main.grid_columnconfigure(0,weight=1); main.grid_columnconfigure(1,weight=2); main.grid_rowconfigure(0,weight=1)

        # lista de checks
        lf=ctk.CTkFrame(main,fg_color=C["surface"],corner_radius=0); lf.grid(row=0,column=0,sticky="nsew"); lf.grid_rowconfigure(1,weight=1); lf.grid_columnconfigure(0,weight=1)
        hf=ctk.CTkFrame(lf,fg_color="transparent"); hf.grid(row=0,column=0,sticky="ew",padx=12,pady=8)
        ctk.CTkLabel(hf,text="Checks",font=ctk.CTkFont(FS,12,"bold"),text_color=C["text"]).pack(side="left")
        ctk.CTkButton(hf,text="+ Novo",width=80,height=28,fg_color=C["accent2"],hover_color="#0255A3",font=ctk.CTkFont(FS,11),command=self._custom_novo).pack(side="right")
        self.custom_list=ctk.CTkScrollableFrame(lf,fg_color="transparent"); self.custom_list.grid(row=1,column=0,sticky="nsew",padx=4)
        self._refresh_custom_list()

        # editor
        ef=ctk.CTkScrollableFrame(main,fg_color=C["panel"],corner_radius=0); ef.grid(row=0,column=1,sticky="nsew"); ef.grid_columnconfigure(0,weight=1)
        ctk.CTkLabel(ef,text="Editor de Check",font=ctk.CTkFont(FS,13,"bold"),text_color=C["text"]).grid(row=0,column=0,padx=16,pady=(16,4),sticky="w")
        ctk.CTkFrame(ef,height=1,fg_color=C["border"]).grid(row=1,column=0,padx=12,sticky="ew")
        self._custom_fields={}
        fields=[("Nome do check","nome","Verificar pedidos atrasados"),("Descrição","descricao","Detecta pedidos sem processamento"),("SQL","sql","SELECT COUNT(*) FROM pedidos WHERE status='PENDENTE' AND created_at < NOW() - INTERVAL 1 DAY"),("Banco alvo","db_tipo","all (ou: sqlite, mysql, postgresql, sqlserver)"),("Comparador","comparador","gt (gt|lt|eq|ne|gte|lte|nonempty|empty)"),("Threshold","threshold","10"),("Status de falha","status_falha","AVISO (ou ERRO)")]
        for i,(label,key,ph) in enumerate(fields):
            ctk.CTkLabel(ef,text=label,font=ctk.CTkFont(FS,11),text_color=C["muted"]).grid(row=i*2+2,column=0,padx=16,pady=(6,0),sticky="w")
            h=80 if key=="sql" else 36
            if key=="sql":
                e=ctk.CTkTextbox(ef,height=h,fg_color=C["surface"],border_color=C["border"],font=ctk.CTkFont(FM,11),border_width=1); e.grid(row=i*2+3,column=0,padx=14,pady=(2,0),sticky="ew")
            else:
                e=ctk.CTkEntry(ef,placeholder_text=ph,height=h,fg_color=C["surface"],border_color=C["border"],font=ctk.CTkFont(FM,11)); e.grid(row=i*2+3,column=0,padx=14,pady=(2,0),sticky="ew")
            self._custom_fields[key]=e

        self._custom_ativo=tk.BooleanVar(value=True)
        ctk.CTkCheckBox(ef,text="Check ativo",variable=self._custom_ativo,font=ctk.CTkFont(FS,12),text_color=C["text2"],fg_color=C["accent2"]).grid(row=len(fields)*2+2,column=0,padx=16,pady=8,sticky="w")
        bf=ctk.CTkFrame(ef,fg_color="transparent"); bf.grid(row=len(fields)*2+3,column=0,padx=14,pady=8,sticky="w")
        ctk.CTkButton(bf,text="💾 Salvar",height=36,width=120,fg_color=C["accent2"],hover_color="#0255A3",font=ctk.CTkFont(FS,12),command=self._custom_salvar).pack(side="left",padx=(0,8))
        ctk.CTkButton(bf,text="🗑 Excluir",height=36,width=100,fg_color=C["danger"],hover_color="#B91C1C",font=ctk.CTkFont(FS,12),command=self._custom_excluir).pack(side="left")

        self._custom_sel_id=None

    def _refresh_custom_list(self):
        for w in self.custom_list.winfo_children(): w.destroy()
        if not self._custom_checks:
            ctk.CTkLabel(self.custom_list,text="Nenhum check.\nClique em '+ Novo'.",font=ctk.CTkFont(FS,11),text_color=C["muted"],justify="center").pack(pady=16); return
        for chk in self._custom_checks:
            card=ctk.CTkFrame(self.custom_list,fg_color=C["surface2"] if chk.ativo else C["surface"],corner_radius=8); card.pack(fill="x",padx=4,pady=2)
            ctk.CTkLabel(card,text=f"{'✅' if chk.ativo else '○'} {chk.nome}",font=ctk.CTkFont(FS,11,"bold"),text_color=C["text"] if chk.ativo else C["muted"]).pack(padx=10,pady=(6,2),anchor="w")
            ctk.CTkLabel(card,text=f"{chk.db_tipo}  •  {chk.comparador} {chk.threshold}  →  {chk.status_falha}",font=ctk.CTkFont(FM,10),text_color=C["muted"]).pack(padx=10,pady=(0,6),anchor="w")
            for w in [card]+card.winfo_children(): w.bind("<Button-1>",lambda e,c=chk:self._custom_selecionar(c))

    def _custom_selecionar(self,chk:CustomCheck):
        self._custom_sel_id=chk.id
        for key,field in self._custom_fields.items():
            val=getattr(chk,key,"")
            if isinstance(field,ctk.CTkTextbox): field.delete("0.0","end"); field.insert("0.0",str(val))
            else: field.delete(0,"end"); field.insert(0,str(val))
        self._custom_ativo.set(chk.ativo)

    def _custom_novo(self):
        self._custom_sel_id=None
        for field in self._custom_fields.values():
            if isinstance(field,ctk.CTkTextbox): field.delete("0.0","end")
            else: field.delete(0,"end")
        self._custom_ativo.set(True)

    def _custom_salvar(self):
        def gv(k):
            f=self._custom_fields[k]
            return f.get("0.0","end").strip() if isinstance(f,ctk.CTkTextbox) else f.get().strip()
        nome=gv("nome")
        if not nome: messagebox.showwarning("Check","Informe o nome do check."); return
        import uuid
        # FIX: usa uuid para novos checks — md5(nome) causava colisão de IDs entre checks com mesmo nome
        chk=CustomCheck(
            id=self._custom_sel_id or uuid.uuid4().hex[:8],
            nome=nome, descricao=gv("descricao"), sql=gv("sql"),
            db_tipo=gv("db_tipo") or "all", comparador=gv("comparador") or "gt",
            threshold=gv("threshold"), status_falha=gv("status_falha") or "AVISO",
            ativo=self._custom_ativo.get()
        )
        existing=[c for c in self._custom_checks if c.id!=chk.id]
        self._custom_checks=existing+[chk]; self._custom_sel_id=chk.id
        self._refresh_custom_list()

    def _custom_excluir(self):
        if not self._custom_sel_id: return
        self._custom_checks=[c for c in self._custom_checks if c.id!=self._custom_sel_id]
        self._custom_sel_id=None; self._custom_novo(); self._refresh_custom_list()

    # ════════════════════════════════════════════════════════════════════════
    # PAGE: NOTIFICAÇÕES
    # ════════════════════════════════════════════════════════════════════════
    def _build_page_notificacoes(self):
        p=self.pages["notificacoes"]; p.grid_columnconfigure(0,weight=1); p.grid_rowconfigure(1,weight=1)
        self._phdr(p,"🔔 Notificações")
        scroll=ctk.CTkScrollableFrame(p,fg_color=C["bg"]); scroll.grid(row=1,column=0,sticky="nsew"); scroll.grid_columnconfigure(0,weight=1)

        def entry(parent,label,key,ph="",show="",row=0,col=0,w=300):
            ctk.CTkLabel(parent,text=label,font=ctk.CTkFont(FS,11),text_color=C["muted"]).grid(row=row*2,column=col,sticky="w",padx=8,pady=(6,0))
            e=ctk.CTkEntry(parent,placeholder_text=ph,show=show,fg_color=C["panel"],border_color=C["border"],font=ctk.CTkFont(FM,12),width=w,height=36)
            e.grid(row=row*2+1,column=col,padx=8,pady=(2,4),sticky="ew"); self._notif_entries[key]=e

        self._notif_entries={}; self._notif_vars={}
        cfg=self._notif_cfg

        # Windows
        wc=self._card(scroll,"🖥 Windows Toast",row=0)
        v=tk.BooleanVar(value=cfg.windows_enabled); self._notif_vars["windows_enabled"]=v
        ctk.CTkCheckBox(wc,text="Habilitar notificações do Windows",variable=v,font=ctk.CTkFont(FS,12),text_color=C["text2"],fg_color=C["accent2"]).grid(row=2,column=0,padx=16,pady=12,sticky="w")

        # Email
        ec=self._card(scroll,"📧 Email (SMTP)",row=1); ec.grid_columnconfigure((0,1),weight=1)
        v2=tk.BooleanVar(value=cfg.email_enabled); self._notif_vars["email_enabled"]=v2
        ctk.CTkCheckBox(ec,text="Habilitar notificação por email",variable=v2,font=ctk.CTkFont(FS,12),text_color=C["text2"],fg_color=C["accent2"]).grid(row=2,column=0,columnspan=2,padx=16,pady=8,sticky="w")
        for lbl,key,ph,show,row,col in [("Servidor SMTP","smtp_host","smtp.gmail.com","",1,0),("Porta","smtp_port","587","",1,1),("Usuário","smtp_user","seu@email.com","",2,0),("Senha","smtp_pass","••••","●",2,1),("De (remetente)","email_from","noreply@empresa.com","",3,0),("Para (destinatários)","email_to","ti@empresa.com, gestor@empresa.com","",3,1)]:
            entry(ec,lbl,key,ph,show,row,col,w=260)
        for k,v in [("smtp_host",cfg.smtp_host),("smtp_port",str(cfg.smtp_port)),("smtp_user",cfg.smtp_user),("email_from",cfg.email_from),("email_to",cfg.email_to)]:
            if v and k in self._notif_entries: self._notif_entries[k].delete(0,"end"); self._notif_entries[k].insert(0,v)

        # Slack
        sc=self._card(scroll,"💬 Slack Webhook",row=2); sc.grid_columnconfigure(0,weight=1)
        v3=tk.BooleanVar(value=cfg.slack_enabled); self._notif_vars["slack_enabled"]=v3
        ctk.CTkCheckBox(sc,text="Habilitar notificação no Slack",variable=v3,font=ctk.CTkFont(FS,12),text_color=C["text2"],fg_color=C["accent2"]).grid(row=2,column=0,padx=16,pady=8,sticky="w")
        ctk.CTkLabel(sc,text="Webhook URL",font=ctk.CTkFont(FS,11),text_color=C["muted"]).grid(row=3,column=0,padx=16,pady=(4,0),sticky="w")
        we=ctk.CTkEntry(sc,placeholder_text="https://hooks.slack.com/services/...",fg_color=C["panel"],border_color=C["border"],font=ctk.CTkFont(FM,12),height=36)
        we.grid(row=4,column=0,padx=14,pady=(2,8),sticky="ew"); self._notif_entries["slack_webhook"]=we
        if cfg.slack_webhook: we.delete(0,"end"); we.insert(0,cfg.slack_webhook)

        # Gatilhos
        gc=self._card(scroll,"⚡ Gatilhos",row=3); gc.grid_columnconfigure((0,1,2),weight=1)
        for i,(k,l) in enumerate([("notify_on_error","Notificar em ERROS"),("notify_on_warning","Notificar em AVISOS"),("notify_on_ok","Notificar em OK")]):
            v=tk.BooleanVar(value=getattr(cfg,k)); self._notif_vars[k]=v
            ctk.CTkCheckBox(gc,text=l,variable=v,font=ctk.CTkFont(FS,12),text_color=C["text2"],fg_color=C["accent2"]).grid(row=2,column=i,padx=16,pady=12,sticky="w")

        # Botões
        bf=ctk.CTkFrame(scroll,fg_color="transparent"); bf.grid(row=4,column=0,pady=12)
        ctk.CTkButton(bf,text="💾 Salvar Configurações",height=42,width=220,fg_color=C["accent2"],hover_color="#0255A3",font=ctk.CTkFont(FS,13,"bold"),command=self._notif_salvar).pack(side="left",padx=(0,10))
        ctk.CTkButton(bf,text="🧪 Testar Canais",height=42,width=160,fg_color=C["surface"],hover_color=C["border"],border_width=1,border_color=C["border"],font=ctk.CTkFont(FS,12),command=self._notif_testar).pack(side="left")

    def _notif_salvar(self):
        def gv(k): return self._notif_entries[k].get().strip() if k in self._notif_entries else ""
        self._notif_cfg=NotifConfig(
            windows_enabled=self._notif_vars.get("windows_enabled",tk.BooleanVar(value=True)).get(),
            email_enabled=self._notif_vars.get("email_enabled",tk.BooleanVar()).get(),
            smtp_host=gv("smtp_host"), smtp_port=int(gv("smtp_port") or 587),
            smtp_user=gv("smtp_user"), smtp_pass=gv("smtp_pass"),
            email_from=gv("email_from"), email_to=gv("email_to"),
            slack_enabled=self._notif_vars.get("slack_enabled",tk.BooleanVar()).get(),
            slack_webhook=gv("slack_webhook"),
            notify_on_error=self._notif_vars.get("notify_on_error",tk.BooleanVar(value=True)).get(),
            notify_on_warning=self._notif_vars.get("notify_on_warning",tk.BooleanVar()).get(),
            notify_on_ok=self._notif_vars.get("notify_on_ok",tk.BooleanVar()).get(),
        )
        self._notif=NotificationDispatcher(self._notif_cfg)
        messagebox.showinfo("Notificações","Configurações salvas!")

    def _notif_testar(self):
        self._notif_salvar()
        results=self._notif.test()
        msg="\n".join(f"{canal}: {res}" for canal,res in results.items()) or "Nenhum canal habilitado."
        messagebox.showinfo("Teste de Notificações",msg)

    # ════════════════════════════════════════════════════════════════════════
    # PAGE: DASHBOARD (com tendência histórica)
    # ════════════════════════════════════════════════════════════════════════
    def _build_page_dashboard(self):
        p=self.pages["dashboard"]; p.grid_columnconfigure(0,weight=1); p.grid_rowconfigure(1,weight=1)
        self._phdr(p,"📊 Dashboard")
        self.dash_scroll=ctk.CTkScrollableFrame(p,fg_color=C["bg"]); self.dash_scroll.grid(row=1,column=0,sticky="nsew"); self.dash_scroll.grid_columnconfigure((0,1,2,3),weight=1)
        ctk.CTkLabel(self.dash_scroll,text="Execute um diagnóstico para ver o dashboard.",font=ctk.CTkFont(FS,13),text_color=C["muted"]).grid(row=0,column=0,columnspan=4,pady=80)

    def _refresh_dashboard(self):
        for w in self.dash_scroll.winfo_children(): w.destroy()
        r=self._resultados
        if not r: return
        er=sum(1 for x in r if x.status=="ERRO"); av=sum(1 for x in r if x.status=="AVISO"); ok=sum(1 for x in r if x.status=="OK"); inf=sum(1 for x in r if x.status=="INFO")
        # KPIs
        for i,(lbl,val,col,sub) in enumerate([("Total",str(len(r)),C["accent"],"checks"),("✅ OK",str(ok),C["ok"],"sem problemas"),("⚠️ Avisos",str(av),C["warning"],"atenção"),("❌ Erros",str(er),C["danger"],"urgente")]):
            kf=ctk.CTkFrame(self.dash_scroll,fg_color=C["surface"],corner_radius=12); kf.grid(row=0,column=i,padx=10,pady=12,sticky="ew")
            ctk.CTkLabel(kf,text=lbl,font=ctk.CTkFont(FS,12),text_color=C["muted"]).pack(padx=16,pady=(14,2),anchor="w")
            ctk.CTkLabel(kf,text=val,font=ctk.CTkFont(FM,36,"bold"),text_color=col).pack(padx=16,anchor="w")
            ctk.CTkLabel(kf,text=sub,font=ctk.CTkFont(FS,11),text_color=C["muted"]).pack(padx=16,pady=(0,14),anchor="w")
        # Donut
        df=ctk.CTkFrame(self.dash_scroll,fg_color=C["surface"],corner_radius=12); df.grid(row=1,column=0,columnspan=2,padx=10,pady=8,sticky="nsew")
        ctk.CTkLabel(df,text="Distribuição por Status",font=ctk.CTkFont(FS,13,"bold"),text_color=C["text"]).pack(padx=16,pady=(12,4),anchor="w")
        fig=Figure(figsize=(4.2,3),facecolor=C["chart_bg"]); ax=fig.add_subplot(111); ax.set_facecolor(C["chart_bg"])
        slices=[(ok,C["ok"],"OK"),(av,C["warning"],"Avisos"),(er,C["danger"],"Erros"),(inf,C["info"],"Info")]
        nz=[(s,c,l) for s,c,l in slices if s>0]
        if nz:
            s_,c_,l_=zip(*nz); _,ts,ats=ax.pie(s_,labels=l_,colors=c_,autopct="%1.0f%%",startangle=90,pctdistance=0.75,wedgeprops=dict(width=0.5,edgecolor=C["chart_bg"],linewidth=2))
            for t in ts: t.set_color(C["chart_fg"]); t.set_fontsize(9)
            for t in ats: t.set_color(C["chart_fg"]); t.set_fontsize(8)
        ax.axis("equal"); fig.tight_layout(); cnv=FigureCanvasTkAgg(fig,master=df); cnv.draw(); cnv.get_tk_widget().pack(fill="both",expand=True,padx=8,pady=(0,12))
        # Bar por categoria
        bf2=ctk.CTkFrame(self.dash_scroll,fg_color=C["surface"],corner_radius=12); bf2.grid(row=1,column=2,columnspan=2,padx=10,pady=8,sticky="nsew")
        ctk.CTkLabel(bf2,text="Por Categoria",font=ctk.CTkFont(FS,13,"bold"),text_color=C["text"]).pack(padx=16,pady=(12,4),anchor="w")
        cats={}
        for x in r: cats.setdefault(x.categoria,{"OK":0,"AVISO":0,"ERRO":0,"INFO":0})[x.status]+=1
        fig2=Figure(figsize=(4.2,3),facecolor=C["chart_bg"]); ax2=fig2.add_subplot(111); ax2.set_facecolor(C["chart_bg"])
        names=list(cats.keys()); xi=range(len(names)); w=0.2
        for j,(st,col) in enumerate([("OK",C["ok"]),("AVISO",C["warning"]),("ERRO",C["danger"]),("INFO",C["info"])]):
            ax2.bar([i+j*w for i in xi],[cats[c].get(st,0) for c in names],w,label=st,color=col,alpha=0.85)
        ax2.set_xticks([i+1.5*w for i in xi]); ax2.set_xticklabels(names,color=C["chart_fg"],fontsize=8); ax2.tick_params(colors=C["chart_fg"]); ax2.spines[:].set_color(C["chart_grid"]); ax2.yaxis.grid(True,color=C["chart_grid"],linewidth=0.5); ax2.legend(fontsize=7,facecolor=C["surface"],edgecolor=C["border"],labelcolor=C["chart_fg"])
        fig2.tight_layout(); cnv2=FigureCanvasTkAgg(fig2,master=bf2); cnv2.draw(); cnv2.get_tk_widget().pack(fill="both",expand=True,padx=8,pady=(0,12))
        # Tendência histórica
        sessions=self._store.list_sessions(20) 
        if len(sessions)>=2:
            tf=ctk.CTkFrame(self.dash_scroll,fg_color=C["surface"],corner_radius=12); tf.grid(row=2,column=0,columnspan=4,padx=10,pady=8,sticky="ew")
            ctk.CTkLabel(tf,text="📈 Tendência Histórica (últimas 20 sessões)",font=ctk.CTkFont(FS,13,"bold"),text_color=C["text"]).pack(padx=16,pady=(12,4),anchor="w")
            sessions_rev=list(reversed(sessions))
            dts=[s["started_at"][11:16] for s in sessions_rev]; erros_h=[s["erros"] for s in sessions_rev]; avisos_h=[s["avisos"] for s in sessions_rev]; ok_h=[s["ok"] for s in sessions_rev]
            fig3=Figure(figsize=(10,2.8),facecolor=C["chart_bg"]); ax3=fig3.add_subplot(111); ax3.set_facecolor(C["chart_bg"])
            xi3=range(len(dts))
            ax3.fill_between(xi3,erros_h,alpha=0.3,color=C["danger"]); ax3.plot(xi3,erros_h,color=C["danger"],label="Erros",linewidth=2,marker="o",markersize=5)
            ax3.fill_between(xi3,avisos_h,alpha=0.2,color=C["warning"]); ax3.plot(xi3,avisos_h,color=C["warning"],label="Avisos",linewidth=2,marker="s",markersize=4)
            ax3.fill_between(xi3,ok_h,alpha=0.1,color=C["ok"]); ax3.plot(xi3,ok_h,color=C["ok"],label="OK",linewidth=1.5,linestyle="--",marker=".",markersize=4)
            ax3.set_xticks(xi3); ax3.set_xticklabels(dts,rotation=35,ha="right",color=C["chart_fg"],fontsize=8); ax3.tick_params(colors=C["chart_fg"]); ax3.spines[:].set_color(C["chart_grid"]); ax3.yaxis.grid(True,color=C["chart_grid"],linewidth=0.5)
            ax3.legend(fontsize=8,facecolor=C["surface"],edgecolor=C["border"],labelcolor=C["chart_fg"])
            fig3.tight_layout(); cnv3=FigureCanvasTkAgg(fig3,master=tf); cnv3.draw(); cnv3.get_tk_widget().pack(fill="x",padx=8,pady=(0,12))
        # Problemas
        prob=[x for x in r if x.status in ("ERRO","AVISO")]
        if prob:
            pf=ctk.CTkFrame(self.dash_scroll,fg_color=C["surface"],corner_radius=12); pf.grid(row=3,column=0,columnspan=4,padx=10,pady=8,sticky="ew")
            ctk.CTkLabel(pf,text="⚡ Problemas Encontrados",font=ctk.CTkFont(FS,13,"bold"),text_color=C["text"]).pack(padx=16,pady=(12,8),anchor="w")
            for x in prob[:8]:
                rf=ctk.CTkFrame(pf,fg_color=C["surface2"],corner_radius=8); rf.pack(fill="x",padx=12,pady=3)
                ctk.CTkLabel(rf,text=f"{ICONS[x.status]} [{x.categoria}] {x.nome}",font=ctk.CTkFont(FM,11,"bold"),text_color=C["danger"] if x.status=="ERRO" else C["warning"]).pack(side="left",padx=12,pady=6)
                ctk.CTkLabel(rf,text=x.detalhe,font=ctk.CTkFont(FS,11),text_color=C["text2"]).pack(side="left",padx=4)
            ctk.CTkFrame(pf,height=12,fg_color="transparent").pack()

    # ════════════════════════════════════════════════════════════════════════
    # PAGES: RESULTADOS, HISTÓRICO, AGENDAMENTO, AUDITORIA (compactos)
    # ════════════════════════════════════════════════════════════════════════
    def _build_page_resultados(self):
        p=self.pages["resultados"]; p.grid_columnconfigure(0,weight=1); p.grid_rowconfigure(2,weight=1)
        self._phdr(p,"📋 Resultados Detalhados")
        fb=ctk.CTkFrame(p,fg_color=C["surface2"],height=44,corner_radius=0); fb.grid(row=1,column=0,sticky="ew"); fb.grid_propagate(False)
        self._filtros_btns={}
        for f in ["Todos","Conexão","Integridade","Performance","Espaço","Custom","ERRO","AVISO"]:
            b=ctk.CTkButton(fb,text=f,width=84,height=28,corner_radius=6,fg_color=C["surface"],hover_color=C["border"],border_width=1,border_color=C["border"],font=ctk.CTkFont(FS,11),command=lambda x=f:self._filtrar(x))
            b.pack(side="left",padx=3,pady=8); self._filtros_btns[f]=b
        tf=ctk.CTkFrame(p,fg_color=C["panel"],corner_radius=0); tf.grid(row=2,column=0,sticky="nsew"); tf.grid_rowconfigure(0,weight=1); tf.grid_columnconfigure(0,weight=1)
        style=ttk.Style(); style.theme_use("clam")
        style.configure("V3.Treeview",background=C["panel"],foreground=C["text"],fieldbackground=C["panel"],rowheight=28,font=(FM,11),relief="flat")
        style.configure("V3.Treeview.Heading",background=C["surface2"],foreground=C["muted"],font=(FS,11,"bold"),relief="flat")
        style.map("V3.Treeview",background=[("selected",C["tree_sel"])],foreground=[("selected",C["text"])])
        self.tree=ttk.Treeview(tf,style="V3.Treeview",columns=("st","cat","nome","det"),show="headings",selectmode="browse")
        for col,hdr,w in [("st","Status",100),("cat","Categoria",120),("nome","Verificação",240),("det","Detalhe",500)]:
            self.tree.heading(col,text=hdr,anchor="w"); self.tree.column(col,width=w)
        for st,col in [("OK",C["ok"]),("ERRO",C["danger"]),("AVISO",C["warning"]),("INFO",C["info"])]: self.tree.tag_configure(st,foreground=col)
        sb2=ctk.CTkScrollbar(tf,command=self.tree.yview); self.tree.configure(yscrollcommand=sb2.set)
        self.tree.grid(row=0,column=0,sticky="nsew"); sb2.grid(row=0,column=1,sticky="ns")
        self.lbl_resumo=ctk.CTkLabel(p,text="Nenhum diagnóstico.",font=ctk.CTkFont(FM,11),text_color=C["muted"]); self.lbl_resumo.grid(row=3,column=0,pady=6)

    def _refresh_resultados(self):
        self._filtrar("Todos")
        r=self._resultados; er=sum(1 for x in r if x.status=="ERRO"); av=sum(1 for x in r if x.status=="AVISO"); ok=sum(1 for x in r if x.status=="OK"); inf=sum(1 for x in r if x.status=="INFO")
        self.lbl_resumo.configure(text=f"Total {len(r)}  •  ✅ {ok}  •  ❌ {er}  •  ⚠️ {av}  •  ℹ️ {inf}",text_color=C["danger"] if er else (C["warning"] if av else C["ok"]))

    def _filtrar(self,filtro):
        for k,b in self._filtros_btns.items(): b.configure(fg_color=C["accent2"] if k==filtro else C["surface"])
        for item in self.tree.get_children(): self.tree.delete(item)
        for r in self._resultados:
            if filtro!="Todos" and filtro not in ("ERRO","AVISO") and r.categoria!=filtro: continue
            if filtro in ("ERRO","AVISO") and r.status!=filtro: continue
            self.tree.insert("","end",values=(f"{ICONS.get(r.status,'')} {r.status}",r.categoria,r.nome,r.detalhe),tags=(r.status,))

    def _build_page_historico(self):
        p=self.pages["historico"]; p.grid_columnconfigure(0,weight=1); p.grid_rowconfigure(1,weight=1)
        self._phdr(p,"🕒 Histórico")
        mh=ctk.CTkFrame(p,fg_color=C["bg"]); mh.grid(row=1,column=0,sticky="nsew"); mh.grid_columnconfigure(0,weight=1); mh.grid_columnconfigure(1,weight=2); mh.grid_rowconfigure(0,weight=1)
        lf=ctk.CTkFrame(mh,fg_color=C["surface"],corner_radius=0); lf.grid(row=0,column=0,sticky="nsew"); lf.grid_rowconfigure(1,weight=1); lf.grid_columnconfigure(0,weight=1)
        ctk.CTkLabel(lf,text="Sessões",font=ctk.CTkFont(FS,12,"bold"),text_color=C["muted"]).grid(row=0,column=0,padx=16,pady=10,sticky="w")
        self.hist_list=ctk.CTkScrollableFrame(lf,fg_color="transparent"); self.hist_list.grid(row=1,column=0,sticky="nsew",padx=4)
        df=ctk.CTkFrame(mh,fg_color=C["panel"],corner_radius=0); df.grid(row=0,column=1,sticky="nsew"); df.grid_rowconfigure(1,weight=1); df.grid_columnconfigure(0,weight=1)
        ctk.CTkLabel(df,text="Detalhe",font=ctk.CTkFont(FS,12,"bold"),text_color=C["muted"]).grid(row=0,column=0,padx=16,pady=10,sticky="w")
        self.hist_tree=ttk.Treeview(df,style="V3.Treeview",columns=("st","cat","nome","det"),show="headings")
        for col,hdr,w in [("st","Status",90),("cat","Categoria",110),("nome","Verificação",200),("det","Detalhe",350)]:
            self.hist_tree.heading(col,text=hdr,anchor="w"); self.hist_tree.column(col,width=w)
        for st,col in [("OK",C["ok"]),("ERRO",C["danger"]),("AVISO",C["warning"]),("INFO",C["info"])]: self.hist_tree.tag_configure(st,foreground=col)
        sbh=ctk.CTkScrollbar(df,command=self.hist_tree.yview); self.hist_tree.configure(yscrollcommand=sbh.set)
        self.hist_tree.grid(row=1,column=0,sticky="nsew"); sbh.grid(row=1,column=1,sticky="ns")

    def _reload_history(self):
        for w in self.hist_list.winfo_children(): w.destroy()
        sessions=self._store.list_sessions(50) 
        if not sessions: ctk.CTkLabel(self.hist_list,text="Nenhum histórico.",font=ctk.CTkFont(FS,11),text_color=C["muted"]).pack(pady=12); return
        for s in sessions:
            dt=s["started_at"][:16].replace("T"," "); er=s["erros"]; av=s["avisos"]; col=C["danger"] if er else (C["warning"] if av else C["ok"])
            card=ctk.CTkFrame(self.hist_list,fg_color=C["surface2"],corner_radius=8); card.pack(fill="x",padx=4,pady=3)
            ctk.CTkLabel(card,text=f"🛢  {s['profile_name']} ({s['db_type']})",font=ctk.CTkFont(FS,11,"bold"),text_color=C["text"]).pack(padx=10,pady=(6,2),anchor="w")
            ctk.CTkLabel(card,text=f"{dt}  •  {s['total']} checks",font=ctk.CTkFont(FM,10),text_color=C["muted"]).pack(padx=10,anchor="w")
            ctk.CTkLabel(card,text=f"❌ {er}  ⚠️ {av}",font=ctk.CTkFont(FS,11),text_color=col).pack(padx=10,pady=(0,6),anchor="w")
            def _load(sid=s["id"]):
                res=self._store.get_session_results(sid)
                for item in self.hist_tree.get_children(): self.hist_tree.delete(item)
                for r in res: self.hist_tree.insert("","end",values=(f"{ICONS.get(r.status,'')} {r.status}",r.categoria,r.nome,r.detalhe),tags=(r.status,))
            for w in [card]+card.winfo_children(): w.bind("<Button-1>",lambda e,fn=_load:fn())

    def _build_page_agendamento(self):
        p=self.pages["agendamento"]; p.grid_columnconfigure(0,weight=1); p.grid_rowconfigure(1,weight=1)
        self._phdr(p,"⚙️ Agendamento")
        scroll=ctk.CTkScrollableFrame(p,fg_color=C["bg"]); scroll.grid(row=1,column=0,sticky="nsew"); scroll.grid_columnconfigure(0,weight=1)
        card=ctk.CTkFrame(scroll,fg_color=C["surface"],corner_radius=12); card.grid(row=0,column=0,padx=20,pady=16,sticky="ew"); card.grid_columnconfigure((0,1),weight=1)
        ctk.CTkLabel(card,text="Agendamento Automático",font=ctk.CTkFont(FS,13,"bold"),text_color=C["text"]).grid(row=0,column=0,columnspan=2,padx=16,pady=(12,4),sticky="w")
        ctk.CTkFrame(card,height=1,fg_color=C["border"]).grid(row=1,column=0,columnspan=2,padx=12,sticky="ew")
        ctk.CTkLabel(card,text="Perfil",font=ctk.CTkFont(FS,11),text_color=C["muted"]).grid(row=2,column=0,padx=16,pady=(8,2),sticky="w")
        profiles=[p.nome for p in (self._store.list_profiles() )]
        self.sched_pv=ctk.StringVar(value=profiles[0] if profiles else "")
        ctk.CTkComboBox(card,values=profiles or ["(nenhum)"],variable=self.sched_pv,fg_color=C["panel"],border_color=C["border"],button_color=C["accent2"],width=220).grid(row=3,column=0,padx=16,pady=(0,12),sticky="w")
        ctk.CTkLabel(card,text="Intervalo (min)",font=ctk.CTkFont(FS,11),text_color=C["muted"]).grid(row=2,column=1,padx=16,pady=(8,2),sticky="w")
        self.sched_interval=ctk.CTkEntry(card,fg_color=C["panel"],border_color=C["border"],font=ctk.CTkFont(FM,12),width=100); self.sched_interval.grid(row=3,column=1,padx=16,pady=(0,12),sticky="w"); self.sched_interval.insert(0,"30")
        bf=ctk.CTkFrame(card,fg_color="transparent"); bf.grid(row=4,column=0,columnspan=2,padx=12,pady=12,sticky="w")
        self.btn_ss=ctk.CTkButton(bf,text="▶  Iniciar",fg_color=C["ok"],hover_color="#059669",height=36,width=150,corner_radius=8,command=self._start_sched); self.btn_ss.pack(side="left",padx=(0,8))
        self.btn_sp=ctk.CTkButton(bf,text="⏹  Parar",fg_color=C["danger"],hover_color="#B91C1C",height=36,width=100,corner_radius=8,state="disabled",command=self._stop_sched); self.btn_sp.pack(side="left")
        lf=ctk.CTkFrame(scroll,fg_color=C["surface"],corner_radius=12); lf.grid(row=1,column=0,padx=20,pady=8,sticky="ew"); lf.grid_columnconfigure(0,weight=1)
        ctk.CTkLabel(lf,text="Log",font=ctk.CTkFont(FS,13,"bold"),text_color=C["text"]).grid(row=0,column=0,padx=16,pady=(12,4),sticky="w")
        ctk.CTkFrame(lf,height=1,fg_color=C["border"]).grid(row=1,column=0,padx=12,sticky="ew")
        self.sched_log=ctk.CTkTextbox(lf,height=180,fg_color=C["panel"],text_color=C["text2"],font=ctk.CTkFont(FM,11),border_width=0)
        self.sched_log.grid(row=2,column=0,padx=12,pady=12,sticky="ew"); self.sched_log.insert("0.0","Nenhuma execução.\n"); self.sched_log.configure(state="disabled")

    def _start_sched(self):
        nome=self.sched_pv.get(); profiles={p.nome:p for p in self._store.list_profiles()}
        if nome not in profiles: messagebox.showerror("Agendamento","Perfil inválido."); return
        try: interval=int(self.sched_interval.get() or 30)
        except: messagebox.showerror("Agendamento","Intervalo inválido."); return
        checks=list(CheckRunner.CHECKS_AVAILABLE.keys())
        cfg=next((p for p in self._store.list_profiles() if p.nome==nome), None)
        self._scheduler.start(cfg,checks,interval,on_done=self._on_sched_done)
        self.btn_ss.configure(state="disabled"); self.btn_sp.configure(state="normal")
        self._sched_log(f"▶ Iniciado — {nome} / {interval} min")

    def _stop_sched(self):
        self._scheduler.stop(); self.btn_ss.configure(state="normal"); self.btn_sp.configure(state="disabled"); self._sched_log("⏹ Parado")

    def _on_sched_done(self,session):
        msg=f"✅ {session.total} checks | ❌ {session.erros} | ⚠️ {session.avisos}"
        self.after(0,lambda:self._sched_log(msg))
        if self._notif:
            level="ERRO" if session.erros else ("AVISO" if session.avisos else "OK")
            self._notif.dispatch(f"Agendamento: {session.profile_name}",f"{session.total} checks | {session.erros} erros | {session.avisos} avisos",level)

    def _sched_log(self,msg):
        ts=datetime.now().strftime("%H:%M:%S"); self.sched_log.configure(state="normal"); self.sched_log.insert("0.0",f"[{ts}] {msg}\n"); self.sched_log.configure(state="disabled")

    def _exportar_excel(self):
        if not self._resultados: messagebox.showinfo("Excel","Execute um diagnóstico primeiro."); return
        path=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],initialfile=f"diagnostico_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill,Font,Alignment
            wb=openpyxl.Workbook(); ws=wb.active; ws.title="Diagnóstico"
            FILLS={"OK":PatternFill("solid",fgColor="1A3A2A"),"ERRO":PatternFill("solid",fgColor="3A1A1A"),"AVISO":PatternFill("solid",fgColor="3A2A0A"),"INFO":PatternFill("solid",fgColor="1A1A3A")}
            FONTS={"OK":Font(color="34D399",bold=True),"ERRO":Font(color="F87171",bold=True),"AVISO":Font(color="FBBF24",bold=True),"INFO":Font(color="818CF8",bold=True)}
            hdrs=["Status","Categoria","Verificação","Detalhe","Timestamp"]
            for c,h in enumerate(hdrs,1):
                cell=ws.cell(row=1,column=c,value=h); cell.font=Font(bold=True,color="E2E8F0"); cell.fill=PatternFill("solid",fgColor="1A2535"); cell.alignment=Alignment(horizontal="center")
            for row_n,r in enumerate(self._resultados,2):
                vals=[r.status,r.categoria,r.nome,r.detalhe,r.ts]
                for c,v in enumerate(vals,1):
                    cell=ws.cell(row=row_n,column=c,value=v)
                    if c==1: cell.font=FONTS.get(r.status,Font()); cell.fill=FILLS.get(r.status,PatternFill())
                    cell.alignment=Alignment(wrap_text=c==4)
            ws.column_dimensions["A"].width=12; ws.column_dimensions["B"].width=16; ws.column_dimensions["C"].width=36; ws.column_dimensions["D"].width=60; ws.column_dimensions["E"].width=22
            # aba de resumo
            ws2=wb.create_sheet("Resumo"); er=sum(1 for r in self._resultados if r.status=="ERRO"); av=sum(1 for r in self._resultados if r.status=="AVISO"); ok=sum(1 for r in self._resultados if r.status=="OK"); inf=sum(1 for r in self._resultados if r.status=="INFO")
            for i,(label,val) in enumerate([("Total",len(self._resultados)),("OK",ok),("Avisos",av),("Erros",er),("Info",inf),("Gerado em",datetime.now().strftime("%d/%m/%Y %H:%M"))],1):
                ws2.cell(row=i,column=1,value=label).font=Font(bold=True); ws2.cell(row=i,column=2,value=val)
            wb.save(path)
            messagebox.showinfo("Excel",f"Exportado: {path}")
        except Exception as e: messagebox.showerror("Excel",str(e))

    def _exportar_html_dlg(self):
        if not self._resultados: messagebox.showinfo("HTML","Execute um diagnóstico primeiro."); return
        path=filedialog.asksaveasfilename(defaultextension=".html",filetypes=[("HTML","*.html")],initialfile=f"diagnostico_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        if not path: return
        BADGE={"OK":"#34D399","ERRO":"#F87171","AVISO":"#FBBF24","INFO":"#818CF8"}
        linhas="".join(f'<tr><td><span class="badge" style="background:{BADGE.get(r.status,"#888")}">{r.status}</span></td><td>{r.categoria}</td><td>{r.nome}</td><td>{r.detalhe}</td></tr>' for r in self._resultados)
        er=sum(1 for r in self._resultados if r.status=="ERRO"); av=sum(1 for r in self._resultados if r.status=="AVISO"); ok=sum(1 for r in self._resultados if r.status=="OK")
        html=f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"><title>DB Diagnóstico v3.0</title><style>*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:Consolas,monospace;background:#0C1016;color:#E2E8F0;padding:40px}}h1{{color:#38BDF8;margin-bottom:4px}}.sub{{color:#475569;margin-bottom:24px;font-size:13px}}.kpis{{display:flex;gap:16px;margin-bottom:28px}}.kpi{{background:#131B24;border-radius:10px;padding:16px 24px}}.kpi .v{{font-size:32px;font-weight:700}}.kpi .l{{font-size:12px;color:#475569}}table{{width:100%;border-collapse:collapse}}th{{background:#1A2535;color:#64748B;text-align:left;padding:10px 14px;font-size:12px}}td{{padding:10px 14px;border-bottom:1px solid #263244;font-size:13px}}tr:hover{{background:#131B24}}.badge{{color:#fff;padding:2px 10px;border-radius:20px;font-size:11px;font-weight:700}}</style></head><body><h1>⬡ DB Diagnóstico v3.0</h1><div class="sub">{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</div><div class="kpis"><div class="kpi"><div class="v" style="color:#38BDF8">{len(self._resultados)}</div><div class="l">Total</div></div><div class="kpi"><div class="v" style="color:#34D399">{ok}</div><div class="l">OK</div></div><div class="kpi"><div class="v" style="color:#F87171">{er}</div><div class="l">Erros</div></div><div class="kpi"><div class="v" style="color:#FBBF24">{av}</div><div class="l">Avisos</div></div></div><table><thead><tr><th>Status</th><th>Categoria</th><th>Verificação</th><th>Detalhe</th></tr></thead><tbody>{linhas}</tbody></table></body></html>"""
        with open(path,"w",encoding="utf-8") as f: f.write(html)
        messagebox.showinfo("HTML",f"Exportado: {path}")

    def _set_status(self,msg): self.lbl_status.configure(text=msg)

if __name__=="__main__":
    app=DBDiagApp()
    app.mainloop()
